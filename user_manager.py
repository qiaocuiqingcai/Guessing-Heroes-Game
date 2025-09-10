import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

class UserManager:
    """用户管理类，负责用户数据的增删改查"""
    
    def __init__(self, data_file='userdata.xlsx'):
        """初始化用户管理器"""
        self.data_file = data_file
        self.init_excel()
    
    def init_excel(self):
        """初始化Excel文件，如果不存在则创建"""
        if not os.path.exists(self.data_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(['id', 'username', 'password', 'total_games', 'win_games', 'lol_games', 'lol_wins', 'wangzhe_games', 'wangzhe_wins', 'last_login'])
            wb.save(self.data_file)
    
    def get_next_id(self):
        """获取下一个可用的用户ID"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        return ws.max_row  # 第一行是标题，所以max_row就是下一个ID
    
    def user_exists(self, username):
        """检查用户名是否已存在"""
        return self.get_user_by_username(username) is not None
    
    def verify_user(self, username, password):
        """验证用户登录信息"""
        user = self.get_user_by_username(username)
        if user and user['password'] == password:
            self.update_user_login(user['id'])
            return True
        return False
    
    def get_user_data(self, username):
        """获取用户数据（不包含密码）"""
        user = self.get_user_by_username(username)
        if user:
            # 移除密码字段
            user.pop('password', None)
        return user
    
    def get_user_by_username(self, username):
        """根据用户名查找用户"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == username:
                return {
                    'id': ws.cell(row=row, column=1).value,
                    'username': ws.cell(row=row, column=2).value,
                    'password': ws.cell(row=row, column=3).value,
                    'total_games': ws.cell(row=row, column=4).value or 0,
                    'win_games': ws.cell(row=row, column=5).value or 0,
                    'lol_games': ws.cell(row=row, column=6).value or 0,
                    'lol_wins': ws.cell(row=row, column=7).value or 0,
                    'wangzhe_games': ws.cell(row=row, column=8).value or 0,
                    'wangzhe_wins': ws.cell(row=row, column=9).value or 0,
                    'last_login': ws.cell(row=row, column=10).value
                }
        return None
    
    def get_user_by_id(self, user_id):
        """根据ID查找用户"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                return {
                    'id': ws.cell(row=row, column=1).value,
                    'username': ws.cell(row=row, column=2).value,
                    'password': ws.cell(row=row, column=3).value,
                    'total_games': ws.cell(row=row, column=4).value or 0,
                    'win_games': ws.cell(row=row, column=5).value or 0,
                    'lol_games': ws.cell(row=row, column=6).value or 0,
                    'lol_wins': ws.cell(row=row, column=7).value or 0,
                    'wangzhe_games': ws.cell(row=row, column=8).value or 0,
                    'wangzhe_wins': ws.cell(row=row, column=9).value or 0,
                    'last_login': ws.cell(row=row, column=10).value
                }
        return None
    
    def get_all_users(self):
        """获取所有用户"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        users = []
        
        for row in range(2, ws.max_row + 1):
            users.append({
                'id': ws.cell(row=row, column=1).value,
                'username': ws.cell(row=row, column=2).value,
                'password': ws.cell(row=row, column=3).value,
                'total_games': ws.cell(row=row, column=4).value or 0,
                'win_games': ws.cell(row=row, column=5).value or 0,
                'lol_games': ws.cell(row=row, column=6).value or 0,
                'lol_wins': ws.cell(row=row, column=7).value or 0,
                'wangzhe_games': ws.cell(row=row, column=8).value or 0,
                'wangzhe_wins': ws.cell(row=row, column=9).value or 0,
                'last_login': ws.cell(row=row, column=10).value
            })
        return users
    
    def add_user(self, username, password):
        """添加新用户"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        user_id = self.get_next_id()
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([user_id, username, password, 0, 0, 0, 0, 0, 0, current_time])
        wb.save(self.data_file)
        
        return user_id
    
    def update_user_login(self, user_id):
        """更新用户登录时间"""
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                ws.cell(row=row, column=10).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                break
        
        wb.save(self.data_file)
    
    def update_game_stats(self, username, game_mode, is_win):
        """更新用户游戏统计信息"""
        user = self.get_user_by_username(username)
        if not user:
            return False
            
        user_id = user['id']
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                # 更新总游戏数
                total_games = ws.cell(row=row, column=4).value or 0
                ws.cell(row=row, column=4).value = total_games + 1
                
                # 更新总胜利数
                if is_win:
                    win_games = ws.cell(row=row, column=5).value or 0
                    ws.cell(row=row, column=5).value = win_games + 1
                
                # 更新特定游戏模式的统计
                if game_mode == 'lol':
                    lol_games = ws.cell(row=row, column=6).value or 0
                    ws.cell(row=row, column=6).value = lol_games + 1
                    
                    if is_win:
                        lol_wins = ws.cell(row=row, column=7).value or 0
                        ws.cell(row=row, column=7).value = lol_wins + 1
                else:  # wangzhe
                    wangzhe_games = ws.cell(row=row, column=8).value or 0
                    ws.cell(row=row, column=8).value = wangzhe_games + 1
                    
                    if is_win:
                        wangzhe_wins = ws.cell(row=row, column=9).value or 0
                        ws.cell(row=row, column=9).value = wangzhe_wins + 1
                
                break
        
        wb.save(self.data_file)
        return True
    
    def clear_user_history(self, username, game_type='all'):
        """清除用户游戏历史记录"""
        user = self.get_user_by_username(username)
        if not user:
            return False
            
        user_id = user['id']
        wb = openpyxl.load_workbook(self.data_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                if game_type == 'all':
                    ws.cell(row=row, column=4).value = 0  # total_games
                    ws.cell(row=row, column=5).value = 0  # win_games
                    ws.cell(row=row, column=6).value = 0  # lol_games
                    ws.cell(row=row, column=7).value = 0  # lol_wins
                    ws.cell(row=row, column=8).value = 0  # wangzhe_games
                    ws.cell(row=row, column=9).value = 0  # wangzhe_wins
                elif game_type == 'lol':
                    # 减去LOL游戏数据
                    total_games = ws.cell(row=row, column=4).value or 0
                    win_games = ws.cell(row=row, column=5).value or 0
                    lol_games = ws.cell(row=row, column=6).value or 0
                    lol_wins = ws.cell(row=row, column=7).value or 0
                    
                    ws.cell(row=row, column=4).value = total_games - lol_games
                    ws.cell(row=row, column=5).value = win_games - lol_wins
                    ws.cell(row=row, column=6).value = 0  # lol_games
                    ws.cell(row=row, column=7).value = 0  # lol_wins
                elif game_type == 'wangzhe':
                    # 减去王者荣耀游戏数据
                    total_games = ws.cell(row=row, column=4).value or 0
                    win_games = ws.cell(row=row, column=5).value or 0
                    wangzhe_games = ws.cell(row=row, column=8).value or 0
                    wangzhe_wins = ws.cell(row=row, column=9).value or 0
                    
                    ws.cell(row=row, column=4).value = total_games - wangzhe_games
                    ws.cell(row=row, column=5).value = win_games - wangzhe_wins
                    ws.cell(row=row, column=8).value = 0  # wangzhe_games
                    ws.cell(row=row, column=9).value = 0  # wangzhe_wins
                
                break
        
        wb.save(self.data_file)
        return True

    def get_leaderboard(self, game_mode='all'):
        """获取排行榜数据
        参数:
            game_mode: 'all' - 全部游戏模式, 'lol' - LOL模式, 'wangzhe' - 王者模式
        返回:
            按(游戏次数^(1/3)) * 胜率 排序的用户列表
        """
        users = self.get_all_users()
        
        ranked_users = []
        for user in users:
            if game_mode == 'lol':
                games = user['lol_games']
                wins = user['lol_wins']
            elif game_mode == 'wangzhe':
                games = user['wangzhe_games']
                wins = user['wangzhe_wins']
            else:  # all
                games = user['total_games']
                wins = user['win_games']
            
            if games > 0:
                win_rate = wins / games
                # 计算排名分数：(游戏次数^(1/3)) * 胜率
                score = (games ** (1/3)) * win_rate
            else:
                score = 0
                win_rate = 0
            
            ranked_users.append({
                'username': user['username'],
                'games': games,
                'wins': wins,
                'win_rate': win_rate,
                'score': score
            })
        
        # 按分数降序排序
        ranked_users.sort(key=lambda x: x['score'], reverse=True)
        
        # 添加排名序号
        for i, user in enumerate(ranked_users, 1):
            user['rank'] = i
        
        return ranked_users