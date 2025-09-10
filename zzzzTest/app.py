from flask import Flask, render_template, request, redirect, url_for, session, flash
import openpyxl
from openpyxl import Workbook
import os
from functools import wraps

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # 生产环境请使用更复杂的密钥

# Excel文件配置
EXCEL_FILE = 'userdata.xlsx'

def init_excel():
    """初始化Excel文件，如果不存在则创建"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(['id', 'username', 'password', 'money'])
        wb.save(EXCEL_FILE)

def get_next_id():
    """获取下一个可用的用户ID"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    return ws.max_row  # 第一行是标题，所以max_row就是下一个ID

def get_user_by_username(username):
    """根据用户名查找用户"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == username:
            return {
                'id': ws.cell(row=row, column=1).value,
                'username': ws.cell(row=row, column=2).value,
                'password': ws.cell(row=row, column=3).value,
                'money': ws.cell(row=row, column=4).value
            }
    return None

def get_user_by_id(user_id):
    """根据ID查找用户"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            return {
                'id': ws.cell(row=row, column=1).value,
                'username': ws.cell(row=row, column=2).value,
                'password': ws.cell(row=row, column=3).value,
                'money': ws.cell(row=row, column=4).value
            }
    return None

def get_all_users():
    """获取所有用户"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    users = []
    
    for row in range(2, ws.max_row + 1):
        users.append({
            'id': ws.cell(row=row, column=1).value,
            'username': ws.cell(row=row, column=2).value,
            'password': ws.cell(row=row, column=3).value,
            'money': ws.cell(row=row, column=4).value
        })
    return users

def update_user_money(user_id, new_money):
    """更新用户余额"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            ws.cell(row=row, column=4).value = new_money
            break
    
    wb.save(EXCEL_FILE)

def add_user(username, password, money=0):
    """添加新用户"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    user_id = get_next_id()
    ws.append([user_id, username, password, money])
    wb.save(EXCEL_FILE)
    
    return user_id

# 装饰器：要求登录
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录！', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 装饰器：要求管理员权限
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录！', 'error')
            return redirect(url_for('login'))
        # 这里假设ID为1的用户是管理员，可以根据需要修改
        if session['user_id'] != 1:
            flash('需要管理员权限！', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# 路由定义
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = get_user_by_username(username)
        if user and user['password'] == password:
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash('登录成功！', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('用户名或密码错误！', 'error')
    elif 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not username or not password:
            flash('用户名和密码不能为空！', 'error')
        elif password != confirm_password:
            flash('两次输入的密码不一致！', 'error')
        elif get_user_by_username(username):
            flash('用户名已存在！', 'error')
        else:
            user_id = add_user(username, password, 100)  # 新用户默认100元
            flash('注册成功！请登录。', 'success')
            return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/dashboard')
@login_required
def dashboard():
    user = get_user_by_id(session['user_id'])
    return render_template('dashboard.html', user=user)

@app.route('/admin')
@admin_required
def admin_panel():
    users = get_all_users()
    return render_template('admin.html', users=users)

@app.route('/update_money', methods=['POST'])
@admin_required
def update_money():
    user_id = request.form.get('user_id')
    new_money = request.form.get('money')
    
    try:
        user_id = int(user_id)
        new_money = float(new_money)
        
        user = get_user_by_id(user_id)
        if user:
            update_user_money(user_id, new_money)
            flash(f'用户 {user["username"]} 的余额已更新为 {new_money} 元', 'success')
        else:
            flash('用户不存在！', 'error')
    except ValueError:
        flash('请输入有效的数字！', 'error')
    
    return redirect(url_for('admin_panel'))

@app.route('/logout')
def logout():
    session.clear()
    flash('您已成功退出登录！', 'success')
    return redirect(url_for('login'))

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

if __name__ == '__main__':
    init_excel()
    app.run(debug=True, host='0.0.0.0', port=5000)